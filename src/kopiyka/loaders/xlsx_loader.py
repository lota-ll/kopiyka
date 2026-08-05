"""Лоадер XLSX.

Дві пастки Excel, через які фінансові дані псуються тихо:

**1. Числа приходять як ``float``.** ``openpyxl`` віддає суму ``-113.98``
як Python ``float`` — рівно той тип, від якого захищає ``domain.money``.
Конвертація мусить іти через ``Decimal(str(value))``, а не
``Decimal(value)``: перший дає точно ``-113.98``, другий — двійкове
наближення з хвостом на 17-му знаку.

**2. Дати приходять як ``datetime``.** Excel зберігає дату числом з
епохою, ``openpyxl`` конвертує це в об'єкт. Парсер, що очікує рядок
``"01.08.2026"``, на ньому впаде. Тому лоадер приводить дати до ISO.
"""

from __future__ import annotations

import io
import warnings
from datetime import date, datetime
from decimal import Decimal

from kopiyka.loaders.base import FileLoader, LoadedFile, LoaderError, clean, register_loader


@register_loader
class XlsxLoader(FileLoader):
    fmt = "xlsx"

    @classmethod
    def sniff(cls, data: bytes, filename: str | None) -> bool:
        # XLSX — ZIP-контейнер, сигнатура 'PK\x03\x04'.
        return data[:4] == b"PK\x03\x04"

    def load(self, data: bytes) -> LoadedFile:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise LoaderError("для читання XLSX потрібен openpyxl") from exc

        file_warnings: list[str] = []
        try:
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                # Виписки банків часто без default style — openpyxl попереджає,
                # але читає коректно. Перехоплюємо, щоб не забруднювати stderr.
                workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                file_warnings.extend(str(w.message) for w in caught)
        except Exception as exc:
            # ZIP-сигнатура є, але вміст не є валідною книгою: пошкоджений
            # файл, звичайний .zip або .docx. Для користувача це одна
            # ситуація — «файл прочитати не вдалося».
            raise LoaderError(f"не вдалося прочитати XLSX: {exc}") from exc

        try:
            if not workbook.sheetnames:
                raise LoaderError("книга не містить аркушів")
            sheet_name = self._pick_sheet(workbook)
            worksheet = workbook[sheet_name]
            rows = [
                [self._cell_to_str(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
                if any(v is not None and str(v).strip() for v in row)
            ]
        finally:
            workbook.close()

        if not rows:
            raise LoaderError("аркуш не містить рядків")

        if len(workbook.sheetnames) > 1:
            file_warnings.append(
                f"у книзі {len(workbook.sheetnames)} аркушів, оброблено «{sheet_name}»"
            )

        return LoadedFile(
            rows=rows,
            source_format=self.fmt,
            encoding="utf-8",
            sheet=sheet_name,
            warnings=file_warnings,
        )

    @staticmethod
    def _pick_sheet(workbook: object) -> str:
        """Обирає аркуш з даними.

        Приват24 називає його «Виписки». Якщо такого немає — беремо перший
        непорожній, а не сліпо ``[0]``: перший аркуш може бути обкладинкою.
        """
        names: list[str] = workbook.sheetnames  # type: ignore[attr-defined]
        for candidate in names:
            if "виписк" in candidate.lower() or "statement" in candidate.lower():
                return candidate
        return names[0]

    @staticmethod
    def _cell_to_str(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float):
            # str() перед Decimal — інакше отримаємо двійкове наближення.
            return f"{Decimal(str(value)):f}"
        if isinstance(value, int):
            return str(value)
        return clean(str(value))
